

import json
import string

from config import FRONTEND_URL
from flask import Flask, render_template, url_for, request, redirect, Response, session, flash, send_from_directory, \
    send_file, abort, jsonify, logging, make_response
from flask_cors import CORS, cross_origin
import glob
import os
import jwt
from userfunctions import toSendEmail, RandomStringwithDigitsAndSymbols,checkToken
from functools import wraps
from encryption import DataEncryption
from openpyxl import *

import pyodbc
import xlsxwriter
from datetime import datetime, timedelta

from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Alignment
import logging
logging.basicConfig(filename=r"C:\easyRenewals\easyRenewals_UAT\UATlog.txt", filemode='a', format='%(message)s')

# conn = pyodbc.connect('Driver={SQL Server};'
#                       'Server=Vwbbdr03dbs04;'
#                       'Database=MantraDB_easyrenewals;'
#                       'Trusted_Connection=Yes;''UID=fnol_admin;''PWD=123456;')

conn = pyodbc.connect('Driver={SQL Server};'
                       'Server=Vwbbdr03dbs04;'
                       'Database=MantraDB_easyrenewals;'
                       'Trusted_Connection=Yes;')






app = Flask(__name__)
CORS(app)

de = DataEncryption()
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

def login_required(f):
    @wraps(f)
    def wrap(*args, **kwargs):
        if 'logged_in' in session:
            return f(*args, **kwargs)
        else:
            flash("You need to login first")
            return redirect(url_for('login'))

    return wrap


# For login admin/user
@app.route('/login', methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def login():
    try:
        print("In LOGIN")
        print(request.data)
        result = False
        # Bytes to dictionary
        data = json.loads(request.data, encoding='utf-8')
        email = data["email"]
        password = data["password"]
        print(email)
        print(password)
        enpwd = de.encrypt(password)
        print(enpwd)
        cursor1 = conn.cursor()
        cursor1.execute("select Name,Permissions,Email,temppass,deleted from users_man where Email = ? and curpassword = ?",(email,enpwd))
        count = len(list(cursor1))
        print(count)
        status = None
        temppass = None
        lastupdatedpassworddatetime = None
        cur1 = conn.cursor()
        cur1.execute("select Name,Permissions,Email,temppass,deleted,lastupdatedpassworddatetime from users_man where Email = ? and curpassword = ?",(email,enpwd))
        for r in cur1:
            status = r[4]
            temppass = r[3]
            lastupdatedpassworddatetime = r[5]
        print(status)
        print(temppass)
        if lastupdatedpassworddatetime == None:
            print("lastupdated is null")
            print(lastupdatedpassworddatetime)
            lastupdatedpassworddatetime = datetime.today()
            print(lastupdatedpassworddatetime)
        else:
            lastupdatedpassworddatetime = lastupdatedpassworddatetime
        currentdate = datetime.today()
        datediff = abs((currentdate - lastupdatedpassworddatetime).days)

        if count > 0 and status == "ACTIVE" and temppass == 0 and datediff <= 29 :
            print("correct user")
            result = True
            cursor = conn.cursor()
            cursor.execute("select Name,Permissions,Email from users_man where email like ? and curpassword like ? and deleted = 'ACTIVE'",(email, enpwd))
            permissions = None
            name = None
            emailid = None
            for r in cursor:
                name = r[0]
                print(name)
                permissions = r[1]
                emailid = r[2]
            current_time = datetime.now().timestamp()
            exp = datetime.utcnow() + timedelta(hours=3, minutes=0, seconds=0)
            token = jwt.encode(
                {"exp": exp, "iat": current_time, "permissions": permissions, "Name": name, "Email": emailid},
                algorithm='HS256', key='secret')
            print(token)
            return jsonify({"permissions": permissions, "email": emailid, "name": name, "Token": token.decode("utf-8")}), 200
        elif temppass == 1 and status == "ACTIVE":
            print("First time login")
            return jsonify({"message": "New user, should navigate to password change page","username" : email}), 200
        elif temppass == 0 and status == "INACTIVE":
            print("Inactive user")
            return jsonify({"message": "Inactive user, please contact the Admin"}),401
        elif datediff > 29 :
            print("User password has expired")
            return jsonify({"message": "User password expired. Should navigate to password change page", "username": email}), 200
        else:
            return jsonify({"message": "please check your email or password"}), 401
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"result": "Error in login"}), 422

#New user creation
@app.route('/createuser', methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
#@login_required
def createuser():
    result = False
    try:
        datas = request.headers.get('Authorization')
        print(datas)
        res = checkToken(request.headers.get('Authorization'))
        print(res)
        if res:
            print("3")
            print(request.data)
            data = json.loads(request.data, encoding='utf-8')
            createdbyemail = data["createdbyemail"]
            email = data["email"]
            cur1 = conn.cursor()
            cur1.execute("select * from users_man where email = ?",(email))
            if(len(list(cur1)) > 0):
                return jsonify({"message": "Email already available"}), 422
            else:
                name = data["name"]
                permissions = ''
                if data["accounttype"] == "Admin":
                    permissions = "Admin"
                elif data["accounttype"] == "Generic":
                    j = 1
                    for i in data["permissions"]:
                        if len(data["permissions"]) == j:
                            permissions = permissions + i
                        else:
                            permissions = permissions + i + ','
                        j = j + 1

                print(permissions)
                createddate = datetime.today().strftime('%Y-%m-%d')
                print(createddate)
                password = RandomStringwithDigitsAndSymbols()
                encryptpas = de.encrypt(password)
                print(encryptpas)
                status = "ACTIVE"
                cursor = conn.cursor()
                cursor.execute(
                    "INSERT INTO Users_man (Name,Permissions,Email,curPassword,Createdby,temppass,createddate,deleted,temppassword) VALUES(?,?,?,?,?,?,?,?,?)",
                    (name, permissions, email, encryptpas, createdbyemail, 1, createddate, status, encryptpas))
                conn.commit()
                subject = "New user creation "
                message = f'<p>Hello {name},</p>' + \
                          f'<p>Your account has been created successfully.</p>' + \
                          f'<p>You can login to the application using the following link and credentials -</p>' + \
                          f'<p><a href="{FRONTEND_URL}"><b>{FRONTEND_URL}</b></a></p>' + \
                          f'<p>User name :<strong> {email}</strong></p>' + \
                          f'<p>Password :<strong> {password}</strong></p>' + \
                          f'<p>Once you login to the application you will be navigated to password change page. Please change the password and relogin.</p>' + \
                          f'<p style="color:Red;">**This is computer generated email, please do not respond to this</p>' + \
                          f'<p>Regards</p>' + \
                          f'<p><strong>ICBL</strong></p>'
                r = toSendEmail(email, subject, message)
                print(r)
                result = True
                return jsonify({"result": result}), 200
        else:
            return jsonify({"result": result}), 401
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "email already exists"}), 422



# To view the users data
@app.route('/usersview', methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def usersview():
    result = False
    try:
        datas = request.headers.get('Authorization')
        print(datas)
        res = checkToken(datas)
        print(res)
        if res:
            datas = datas.replace('"','')
            print(datas)
            decoded = jwt.decode(datas, str('secret'), 'utf-8')
            print(decoded)
            result = True
            email = decoded["Email"]
            cursor = conn.cursor()
            cursor.execute("select Name,Permissions,Email,deleted from users_man where email not like ? and permissions not like 'SAdmin'",email)
            row_headers = [x[0] for x in cursor.description]
            records = cursor.fetchall()
            json_data = []
            for record in records:
                json_data.append(dict(zip(row_headers, record)))
            return jsonify(json_data), 200
        else:
            return jsonify({"result": result}), 401
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        result = False
        return jsonify({"result": result}), 422


#permissiosn to frontend
@app.route('/permissionsforui', methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def permissionsforui():
    try:
        print(request.data)
        data = json.loads(request.data, encoding='utf-8')
        email = data["email"]
        print(email)
        cursor = conn.cursor()
        cursor.execute("select permissions from users_man where email like ?",(email))
        json_data = []
        perm = None
        for record in cursor:
            perm = record[0] + ','
        print(perm)
        substr = ''
        for element in range(0, len(perm)):
            if perm[element] == ',':
                print(substr)
                json_data.append(substr)
                substr = ''
            else:
                substr = substr + perm[element]
        print(json_data)
        return jsonify(json_data), 200
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "email already exists"}), 422


#Updating the user's name or permissions
@app.route('/updateuser', methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def updateuser():
    result = False
    try:
        datas = request.headers.get('Authorization')
        print(datas)
        res = checkToken(datas)
        print(res)
        if res:
            print(request.data)
            data = json.loads(request.data, encoding='utf-8')
            email = data["email"]
            name = data["name"]
            perm = data["permissions"]
            print(perm)
            permissions = ''
            i = 0
            for i in range(len(perm)):
                if i == len(perm) - 1:
                    permissions = permissions + perm[i]
                else:
                    permissions = permissions + perm[i] + ","
                i = i + 1
            print(permissions)
            cursor = conn.cursor()
            oldname  = None
            cursor.execute("select name from users_man where email like ?",(email))
            r = cursor.fetchone()
            print(r)
            for rcrd in r:
                oldname = rcrd
            print("old name is")
            print(oldname)

            cursor.execute("update users_man set name = ?, permissions = ?,lastupdated = ? where email like ?",(name, permissions, datetime.now(), email))
            cursor.commit()
            cursor.execute("update uw1 set assigneduser = ? where assigneduser like ?",(name,oldname))
            cursor.commit()
            cursor.execute("update financecalltable set assigneduser = ? where assigneduser like ?",(name,oldname))
            cursor.commit()
            result = True
            return jsonify({"result": result}), 200
        else:
            return jsonify({"result": result}), 401
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured while updating the user"}), 422



#Updating the user's status
@app.route('/statuschange', methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def statuschange():
    result = False
    try:
        datas = request.headers.get('Authorization')
        print(datas)
        res = checkToken(datas)
        print(res)
        if res:
            print(request.data)
            data = json.loads(request.data, encoding='utf-8')
            email = data["email"]
            status = data["status"]
            print(email)
            print(status)

            # if status == 'ACTIVE':
            #     status = 'Active'

            cursor = conn.cursor()
            cursor.execute("update users_man set deleted = ?, lastupdated = ? where email like ?",
            (status, datetime.now(), email))
            cursor.commit()
            result = True
            return jsonify({"result": result}), 200
        else:
            return jsonify({"result": result}), 401
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured while updating the user status"}), 422


# Deleting user from the users list
@app.route('/deleteuser', methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def deleteuser():
    result = False
    try:
        datas = request.headers.get('Authorization')
        print(datas)
        res = checkToken(request.headers.get('Authorization'))
        print(res)
        if res:
            print(request.data)
            data = json.loads(request.data, encoding='utf-8')
            email = data["email"]
            cursor=conn.cursor()
            cursor.execute("delete from users_man where email like ?",(email))
            cursor.commit()
            result = True
            return jsonify({"result": result}), 200
        else:
            return jsonify({"result": result}), 401
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        conn.rollback()
        return jsonify({"message": "Error occured while deleting the user"}), 422

#Sending Temperory password to the Mail for Forgot password users or Temporary password users
@app.route('/tempuserpasswordchange', methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def tempuserpasswordchange():
    try:
        print(request.data)
        result = False
        data = json.loads(request.data, encoding='utf-8')
        email = data["email"]
        pwd = data["newPassword"]
        temppwd = data["oldPassword"]
        cursor = conn.cursor()
        try:

            lastupdatedpassworddatetime = None
            cur1 = conn.cursor()
            cur1.execute(
                "select lastupdatedpassworddatetime from users_man where Email = ?",(email))
            for r in cur1:
                lastupdatedpassworddatetime = r[0]

            if lastupdatedpassworddatetime == None:
                lastupdatedpassworddatetime = datetime.today()
                print(lastupdatedpassworddatetime)
            else:
                lastupdatedpassworddatetime = lastupdatedpassworddatetime
            currentdate = datetime.today()
            datediff = abs((currentdate - lastupdatedpassworddatetime).days)

            if datediff > 29 :
                print("Password has expired")
                cursor.execute("select * from users_man where email like ? and curpassword = ?", (email,de.encrypt(temppwd)))
            else:
                cursor.execute("select * from users_man where email like ? and temppassword = ?", (email,de.encrypt(temppwd)))
            if (len(list(cursor))) > 0:
                cursor1 = conn.cursor()
                cursor1.execute("update users_man set curpassword = ?,temppass = ?,lastupdatedpassworddatetime = ? where email = ?", (de.encrypt(pwd),0,datetime.today(),email))
                cursor1.commit()
                result = True
                return jsonify({"result": result}), 200
            else:
                return jsonify({"message": "User or temppassword is incorrect"}), 422
        except Exception as e:
            write_to_file(e)
            return jsonify({"message": "Error occured while changing the password"}), 422
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in temporary password change"}), 422


#Sending Temperory password to the Mail for Forgot password users
@app.route('/passwordrequest', methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def passwordrequest():
    try:
        print(request.data)
        data = json.loads(request.data, encoding='utf-8')
        email = data["Email"]
        cursor = conn.cursor()
        try:
            cursor.execute("select * from users_man where email like ?", (email))
            if (len(list(cursor))) > 0:
                cursor1 = conn.cursor()
                cursor1.execute("select name from users_man where email like ?", (email))
                name = None
                for r in cursor1:
                    name = r[0]
                print(name)
                password = RandomStringwithDigitsAndSymbols()
                print(password)
                subject = "Password Reset"
                print(subject)
                message = f'<p>Hello {name},</p>' + \
                          f'<p>We received a request to reset your password.</p>' + \
                          f'<p>Please find the below computer-generated password to further reset your password.</p>' + \
                          f'<p>Password : <strong>{password}</strong></p>' + \
                          f'<p>Once you enter the above password you will be navigated to password change page. Please change the password and relogin.</p>' + \
                          f'<p style="color:Red;">**This is computer generated email, please do not respond to this</p>' + \
                          f'<p>Regards</p>' + \
                          f'<p><strong>ICBL</strong></p>'
                r = toSendEmail(email, subject, message)
                print(r)
                encryptedpwd = de.encrypt(password)
                cursor1 = conn.cursor()
                cursor1.execute("update users_man set temppassword = ?,curpassword =?, temppass = ? where email like ?",(encryptedpwd,encryptedpwd,1,email))
                cursor1.commit()
                result = True
                return jsonify({"result": result}), 200
            else:
                result = False
                return jsonify({"message": "no user with the provided mail id"}), 422
        except Exception as e:
            write_to_file(e)
            return jsonify({"message": "Error occured while sending the Temporary password"}), 422
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in password request"}), 422



# For Underwriter Module
@app.route('/uw', methods=['GET','POST','OPTIONS'])
@cross_origin(origins="*", headers="*")
def uw():
    try:
        cur = conn.cursor()
        res = checkToken(request.headers.get('Authorization'))
        print(res)
        result = False
        if res:
            print("in UW")
            data = request.json
            print(data)

            datas = request.headers.get('Authorization')
            # print("-----------------",datetime.now())
            print(datas)
            datas = datas.replace('"', '')
            print(datas)
            decoded = jwt.decode(datas, str('secret'), 'utf-8')
            user = decoded.get("Name")
            print(user)
            permissions = decoded.get("permissions")
            print(permissions)

            try:

                cur.execute("select distinct uw1.*,cast(getdate() as date) as todaydate,cast(uw1.dateofaction as date) as completeddate from uw1 where uw1.assigneduser = ?",(user))

                row_headers = [x[0] for x in cur.description]
                records = cur.fetchall()
                json_data = []
                for record in records:
                    json_data.append(dict(zip(row_headers, record)))
                print("Data is")
                print(json_data)
                result = True
                return jsonify(json_data)
            except jwt.ExpiredSignatureError:
                print("sign")
                result = False
            if result:
                return jsonify({"result": result}), 200
            else:
                return jsonify({"result": result}), 401
        else:
            result = False
            return jsonify({"result": result}), 401
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in UW Module"}), 422



@app.route('/uwclaimhistory', methods=['GET','POST','OPTIONS'])
@cross_origin(origins="*", headers="*")
def uwclaimhistory():
    try:
        cur = conn.cursor()
        res = checkToken(request.headers.get('Authorization'))
        print(res)
        result = False
        if res:
            data = request.json
            print(data)
            polnum = data['policynum']
            print(polnum)
            polnum = polnum.strip()

            try:
                print(polnum)
                cur.execute("select * from claimhistorytable where policynum = ? order by claimnum", polnum)
                if len(list(cur)) == 0:
                    print("no claim data")
                    cur1 = conn.cursor()
                    print(polnum)
                    cur1.execute("select policynum,policyholder,product from uw1 where policynum = ?", polnum)
                    row_headers = [x[0] for x in cur1.description]
                    records = cur1.fetchall()
                    print(records)
                    json_data = []
                    for record in records:
                        json_data.append(dict(zip(row_headers, record)))
                        print("in for")
                    print("Data is")
                    print(json_data)
                    result = True
                    lossratio = 0
                    return jsonify({"jsondata": json_data, "lossratio": lossratio})
                else:
                    cur.execute("select * from claimhistorytable where policynum = ? order by claimnum", polnum)
                    row_headers = [x[0] for x in cur.description]
                    records = cur.fetchall()
                    claimdetails = []
                    for record in records:
                        claimdetails.append(dict(zip(row_headers, record)))
                    print("Data is")
                    print(claimdetails)
                    result = True
                    cur2 = conn.cursor()
                    cur2.execute("select lossratio from polamountslookup where policynum = ?", polnum)
                    lossratio = 0
                    r = cur2.fetchall()
                    for rcd in r:
                        lossratio = rcd[0]
                    print(lossratio)
                    cur1 = conn.cursor()
                    cur1.execute("select policynum,policyholder,product from uw1 where policynum = ?", polnum)
                    row_heads = [x[0] for x in cur1.description]
                    rcrds = cur1.fetchall()
                    print(records)
                    json_data = []
                    for rr in rcrds:
                        json_data.append(dict(zip(row_heads, rr)))
                        print("in for")
                    print("Data is")
                    print(json_data)
                    return jsonify({"jsondata": json_data, "lossratio": lossratio, "claimdetails": claimdetails})
            except jwt.ExpiredSignatureError:
                print("sign")
                result = False
            if result:
                return jsonify({"result": result}), 200
            else:
                return jsonify({"result": result}), 401
        else:
            result = False
            return jsonify({"result": result}), 401
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in claim history"}), 422


@app.route('/finance', methods=['GET','POST','OPTIONS'])
@cross_origin(origins="*", headers="*")
def finance():
    try:
        cur = conn.cursor()
        res = checkToken(request.headers.get('Authorization'))
        print(res)
        result = False
        if res:
            datas = request.headers.get('Authorization')
            # print("-----------------",datetime.now())
            print(datas)
            datas = datas.replace('"', '')
            print(datas)
            decoded = jwt.decode(datas, str('secret'), 'utf-8')

            user = decoded.get("Name")
            print(user)
            permissions = decoded.get("permissions")
            print(permissions)
            if ("FO" == permissions or "FO,FRP" == permissions):
                cur.execute("select distinct finance1.*,cast(getdate() as date) as todaydate,cast(finance1.dateofaction as date) as completeddate from   finance1 where ((uwreview = 1) or policynum not in (select c.policynum from finance1 f join uw1 c on   c.policynum = f.policynum where c.assigneduser is not null)) and finance1.assigneduser = ?",user)
            row_headers = [x[0] for x in cur.description]
            records = cur.fetchall()
            json_data = []
            for record in records:
                json_data.append(dict(zip(row_headers, record)))
            print("Data is")
            print(json_data)
            result = True
            return jsonify(json_data)
        else:
            return jsonify({"result": result}), 401
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in finance function"}), 422



@app.route('/finc', methods=['GET','POST','OPTIONS'])
@cross_origin(origins="*", headers="*")
def finc():
    try:
        cur = conn.cursor()
        res = checkToken(request.headers.get('Authorization'))
        print(res)
        result = False
        if res:
            data = request.json
            print(data)
            polnum = data['policynum']
            print(polnum)
            cur.execute("select distinct policyperiod as policyperiod,previouspaymentplan as previouspaymentplan,premium as premium,premdue as premdue  from finance1  where policynum like ?",
                (polnum))
            row_headers = [x[0] for x in cur.description]
            records = cur.fetchall()
            jsondata = []
            for record in records:
                jsondata.append(dict(zip(row_headers, record)))
            print("json data Data is")
            print(jsondata)
            cursor = conn.cursor()
            cursor.execute("select distinct PolicyNum,product,PolicyHolder  from finance1  where policynum like ?",(polnum))
            row_header = [x[0] for x in cursor.description]
            records = cursor.fetchall()
            basic_data = []
            for record in records:
                basic_data.append(dict(zip(row_header, record)))
            print("basic Data is")
            print(basic_data)

            result = True
            return jsonify({"json_data": jsondata, "basic_data": basic_data})
        else:
            return jsonify({"result": result}), 401
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in finc function"}), 422

@app.route('/finpaymenthistory', methods=['GET','POST','OPTIONS'])
@cross_origin(origins="*", headers="*")
def finpaymenthistory():
    try:
        cur = conn.cursor()
        res = checkToken(request.headers.get('Authorization'))
        print(res)
        result = False
        if res:
            data = request.json
            print(data)
            polnum = data['policynum']
            print(polnum)
            cur.execute("select distinct * from financepaymenthistory fh where fh.policynum like ?", (polnum))
            row_headers = [x[0] for x in cur.description]
            records = cur.fetchall()
            json_data = []
            for record in records:
                json_data.append(dict(zip(row_headers, record)))
            print("Data is")
            print(json_data)
            result = True
            return jsonify(json_data)
        else:
            return jsonify({"result": result}), 401
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in finance payment history"}), 422


@app.route("/uwaction", methods=['GET','POST','OPTIONS'])
@cross_origin(origins="*", headers="*")
def uwaction():
    try:
        res = checkToken(request.headers.get('Authorization'))
        print(res)
        result = False
        if res:
            Data = request.get_json()
            print(Data)
            polnum = Data.get('policynum')
            print(polnum)
            # actionperformedon = datetime.strptime(datetime.now(), '%m-%d-%Y %H:%M')
            actionperformedon = datetime.today()
            print(actionperformedon)
            typeofaction = Data.get('Action')
            print(typeofaction)
            status = Data.get('status')
            print(status)
            name = Data.get('name')
            print(name)
            cur = conn.cursor()
            if typeofaction == 'Accept':
                updatetype = Data.get('Updatetype')
                print("in if")
                print(updatetype)
                notes = Data.get('updatenotes')
                print(notes)
                cur.execute(
                    "update uw1 set typeofaction = ?, updatetype = ?,note= ?, dateofaction = ?,status = ?,username = ? where PolicyNum like ?",
                    (typeofaction, updatetype, notes, actionperformedon, status, name, polnum))
                cur.commit()
                cur1 = conn.cursor()
                cur1.execute(
                    "update finance1 set uwreview = 1, ReviewDueIn = datediff(d,cast(getdate() as date),dateadd(d,10,cast(getdate() as date))) ,reviewduedate = dateadd(d,10,cast(getdate() as date)) ,createdDate = cast(getdate() as date) where policynum = ?",
                    (polnum))
                cur1.commit()

            else:
                notes = Data.get('Rejectionreason')
                print(notes)
                cur.execute(
                    "update uw1 set typeofaction = ?, note= ?, dateofaction = ?,status = ?,username = ? where PolicyNum like ?",
                    (typeofaction, notes, actionperformedon, status, name, polnum))
                cur.commit()

            return jsonify(Data), 200
        else:
            return jsonify({"result": result}), 401
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in uw action"}), 422



@app.route("/finaction", methods=['GET','POST','OPTIONS'])
@cross_origin(origins="*", headers="*")
def finaction():
    try:
        res = checkToken(request.headers.get('Authorization'))
        print(res)
        result = False
        if res:
            Data = request.get_json()
            print(Data)
            polnum = Data.get('policynum')
            polnum = polnum.strip()
            print(polnum)
            previousplanflag = Data.get('Action')
            newpaymentplan = Data.get('newPaymentplan')
            if newpaymentplan == '':
                newpaymentplan = Data.get('Previousplan')
            print(newpaymentplan)
            status = Data.get('status')
            username = Data.get('name')
            dateofaction = datetime.today()
            cur1 = conn.cursor()
            cur1.execute("insert into icbl_rpa values(?,?,?)", (polnum, newpaymentplan, dateofaction))
            cur1.commit()
            cur = conn.cursor()
            cur.execute(
                "update finance1 set previousplanflag = ?,newpaymentplan = ?,dateofaction = ?,username= ?,status = ? where policynum = ?",
                (previousplanflag, newpaymentplan, dateofaction, username, status, polnum))
            cur.commit()

            return jsonify(Data), 200

        else:
            return jsonify({"result": result}), 401
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in finance action"}), 422


#Needs to be updated - updated
@app.route("/userdashboarddata", methods=['GET','POST','OPTIONS'])
@cross_origin(origins="*", headers="*")
def userdashboarddata():
    try:
        res = checkToken(request.headers.get('Authorization'))
        print(res)
        if res:
            datas = request.headers.get('Authorization').replace('"', '')
            print(datas)
            decoded = jwt.decode(datas, str('secret'), 'utf-8')
            print(decoded)
            permissions = decoded["permissions"]
            print(permissions)
            print(decoded["Name"])
            cur = conn.cursor()
            if permissions == 'SUW,UWRP' or permissions == 'SUW' or permissions == 'UW' or permissions == 'UW,UWRP':
                cur.execute("select pol.AssignedUser,max(pol.tp) as totalpolicies,max(pol.cp) as completedpolicies,max(pol.op) as overduepolicies, max(pol.ep) as escalationpolicies  from ( Select Assigneduser,count(policynum) tp,0 cp,0 op,0 ep from uw1  where assigneduser is not null and status is null and reviewduein between 3 and 6 group by AssignedUser  union  Select Assigneduser,0 tp,count(policynum) cp,0 op,0 ep from uw1  where assigneduser is not null and status = 'completed'  and cast(Dateofaction as date)=cast(getdate() as date) group by AssignedUser  union  Select Assigneduser,0 tp,0 cp,count(policynum) op,0 ep from uw1  where assigneduser is not null and status is null  and ReviewDuein < 0   group by AssignedUser   union  Select Assigneduser,0 tp,0 cp,0 op,count(policynum) ep from uw1  where assigneduser is not null and status is null and ReviewDuein in (0,1,2) group by AssignedUser  ) pol join users_man u  on u.name = pol.AssignedUser  and u.name = ?  group by pol.AssignedUser",(decoded["Name"]))
            elif permissions == 'FO,FRP' or permissions == 'FO':
                cur.execute("select pol.AssignedUser,max(pol.tp) as totalpolicies,max(pol.cp) as completedpolicies,max(pol.op) as overduepolicies, max(pol.ep) as escalationpolicies  from ( Select Assigneduser,count(policynum) tp,0 cp,0 op,0 ep from finance1  where assigneduser is not null and status is null and reviewduein between 3 and 10 and policynum in (select distinct policynum from finance1 where ((uwreview = 1) or policynum not in (select c.policynum from finance1 f join uw1 c on c.policynum = f.policynum where c.assigneduser is not null)))  group by AssignedUser  union  Select Assigneduser,0 tp,count(policynum) cp,0 op,0 ep from finance1  where assigneduser is not null and status = 'completed'  and cast(Dateofaction as date)=cast(getdate() as date) and policynum in (select distinct policynum from finance1 where ((uwreview = 1) or policynum not in (select c.policynum from finance1 f join uw1 c on c.policynum = f.policynum where c.assigneduser is not null)))  group by AssignedUser  union  Select Assigneduser,0 tp,0 cp,count(policynum) op,0 ep from finance1  where assigneduser is not null and status is null  and ReviewDuein < 0 and policynum in (select distinct policynum from finance1 where ((uwreview = 1) or policynum not in (select c.policynum from finance1 f join uw1 c on c.policynum = f.policynum where c.assigneduser is not null)))  group by AssignedUser   union  Select Assigneduser,0 tp,0 cp,0 op,count(policynum) ep from finance1  where assigneduser is not null and status is null and ReviewDuein in (0,1,2) and policynum in (select distinct policynum from finance1 where ((uwreview = 1) or policynum not in (select c.policynum from finance1 f join uw1 c on c.policynum = f.policynum where c.assigneduser is not null)))  group by AssignedUser  ) pol join users_man u  on u.name = pol.AssignedUser  and u.name = ?  group by pol.AssignedUser",(decoded["Name"]))
            elif permissions == 'UWRP':
                cur.execute("select max(pol.tp) as totalpolicies,max(pol.cp) as completedpolicies,max(pol.op) as overduepolicies, max(pol.ep) as escalationpolicies from (Select count(policynum) tp,0 cp,0 op,0 ep from uw1  where assigneduser is not null and reviewduein between 3 and 6 and status is null union Select 0 tp,count(policynum) cp,0 op,0 ep from uw1  where assigneduser is not null and status = 'completed' and cast(Dateofaction as date)=cast(getdate() as date) union Select 0 tp,0 cp,count(policynum) op,0 ep from uw1  where assigneduser is not null and status is null and ReviewDuein < 0  union Select 0 tp,0 cp,0 op,count(policynum) ep from uw1 where assigneduser is not null and status is null and ReviewDuein in (0,1,2))pol")
            elif permissions == 'FRP':
                cur.execute("select max(pol.tp) as totalpolicies,max(pol.cp) as completedpolicies,max(pol.op) as overduepolicies, max(pol.ep) as escalationpolicies from (Select count(policynum) tp,0 cp,0 op,0 ep from finance1  where assigneduser is not null and reviewduein between 3 and 10 and status is null and policynum in (select distinct policynum from finance1 where ((uwreview = 1) or policynum not in (select c.policynum from finance1 f join uw1 c on c.policynum = f.policynum where c.assigneduser is not null))) union Select 0 tp,count(policynum) cp,0 op,0 ep from finance1  where assigneduser is not null and status = 'completed' and cast(Dateofaction as date)=cast(getdate() as date) and policynum in (select distinct policynum from finance1 where ((uwreview = 1) or policynum not in (select c.policynum from finance1 f join uw1 c on c.policynum = f.policynum where c.assigneduser is not null))) union Select 0 tp,0 cp,count(policynum) op,0 ep from finance1  where assigneduser is not null and status is null and ReviewDuein < 0 and policynum in (select distinct policynum from finance1 where ((uwreview = 1) or policynum not in (select c.policynum from finance1 f join uw1 c on c.policynum = f.policynum where c.assigneduser is not null))) union Select count(policynum) tp,0 cp,0 op,count(policynum) ep from finance1 where assigneduser is not null and status is null and ReviewDuein in (0,1,2) and policynum in (select distinct policynum from finance1 where ((uwreview = 1) or policynum not in (select c.policynum from finance1 f join uw1 c on c.policynum = f.policynum where c.assigneduser is not null))))pol")
            row_header = [x[0] for x in cur.description]
            records = cur.fetchall()
            data = []
            for record in records:
                data.append(dict(zip(row_header, record)))
            print("Data is")
            print(data)

            return jsonify(data), 200
        else:
            return jsonify({"result": "result"}), 401
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in user dashbaord function"}), 422

#To retrieve all the policies which are assigned to the users
@app.route("/masterData", methods=['GET','POST','OPTIONS'])
@cross_origin(origins="*", headers="*")
def masterData():
    try:
        res = checkToken(request.headers.get('Authorization'))
        print(res)
        if res:
            cur = conn.cursor()
            cur.execute("select policynum as policynumber,policyholder,expirydate,createddate as dateallocated,cast(reviewduein as varchar(5)) reviewduein,assigneduser as allocatedto from uw1 where assigneduser is not null and status is null")
            row_header = [x[0] for x in cur.description]
            records = cur.fetchall()
            data = []
            for record in records:
                data.append(dict(zip(row_header, record)))
            print("Master data is")
            print(data)

            return jsonify(data), 200
        else:
            return jsonify({"result": "result"}), 401
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in master data"}), 422


#Renewal updates #needs to be updated
@app.route("/usersAndPermissionsForUI", methods=['GET','POST','OPTIONS'])
@cross_origin(origins="*", headers="*")
def usersAndPermissionsForUI():
    try:
        print("in users and permissions")
        res = checkToken(request.headers.get('Authorization'))
        print(res)
        if res:
            cur = conn.cursor()
            data = request.json
            print("data is ")
            print(data)
            perm = data.get('permissions')
            datas = []
            dictionary = {}
            if perm == 'outbound':
                roles = ['OBT','OBM']
            else:
                roles = ['SUW', 'UW']
            print(roles)
            j = 0
            role = ['role', 'role']
            for i in range(len(role)):
                print(roles[j])
                dictionary[role[i]] = roles[j]
                datas.append(dictionary)
                dictionary = {}
                j = j + 1
            k = 0
            j = 1
            for i in range(len(role)):
                datas[k]['id'] = j
                k = k + 1
                j = j + 1

            print(datas)
            j = 0
            for per in datas:
                print("IN for loop of per in datas")
                print(per['role'])
                if per['role'] == 'SUW' or per['role'] == 'UW':
                    print("in inner loop")
                    if per['role'] == 'SUW':
                        cur.execute("select distinct name from users_man where permissions like '%SUW%'")
                    elif per['role'] == 'UW':
                        cur.execute("select distinct name from users_man where permissions in  ('UW','UW,UWRP')")
                else:
                    cur.execute("select distinct name from users_man where permissions like ?", '%' + per['role'] + '%')
                records = cur.fetchall()
                print(records)
                namelist = []
                for i in records:
                    namelist.append(i[0])
                datas[j]['users'] = namelist
                j = j + 1
            print(datas)

            return jsonify(datas), 200
        else:
            return jsonify({"result": "result"}), 401
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in sending permissions to UI"}), 422

#Renewals code updated on 20th jan 2021 #needs to be updated
@app.route("/ChiefuwPolicyaAssignment", methods=['GET','POST','OPTIONS'])
@cross_origin(origins="*", headers="*")
def ChiefuwPolicyaAssignment():
    try:
        res = checkToken(request.headers.get('Authorization'))
        print(res)
        if res:
            Data = request.get_json()
            print(Data)
            print(Data.get('permissions'))
            polnum = Data.get('Policynumbers')
            print(polnum)
            allocatedto = Data['values']['username']
            print(allocatedto)
            allocateddate = datetime.today()
            allocatedby = Data.get('username')
            print(allocatedby)

            if Data.get('permissions') == 'outbound':
                for policynum in polnum:
                    cur = conn.cursor()

                    cur.execute("select assigneduser from financecalltable where policynum like ?",(policynum))
                    r = cur.fetchone()
                    user =None
                    for i in r:
                        user = r[0]
                    print(user)
                    cur.execute("update financecalltable set previouslyallocatedto = ? where policynum like ?",(user,policynum))
                    cur.commit()
                    cur.execute(
                        "update financecalltable set allocateddate = ?,assigneduser= ?, allocatedby = ?  where policynum like ?",
                        (allocateddate, allocatedto, allocatedby, policynum))
                    cur.commit()
            elif Data.get('permissions') == 'uw':
                for policynum in polnum:
                    cur = conn.cursor()

                    cur.execute("select assigneduser from uw1 where policynum like ?",(policynum))
                    r = cur.fetchone()
                    user =None
                    for i in r:
                        user = r[0]
                    print(user)
                    cur.execute("update uw1 set previouslyallocatedto = ? where policynum like ?",(user,policynum))
                    cur.commit()
                    cur.execute(
                        "update uw1 set allocateddate = ?,assigneduser= ?, allocatedby = ?  where policynum like ?",
                        (allocateddate, allocatedto, allocatedby, policynum))
                    cur.commit()

            return jsonify({"result":"success"}), 200

        else:
            return jsonify({"result": "result"}), 401
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in  policy Assignment function"}), 422

#updates 20th jan 2021 #needs to be updated

#To retrieve all the policies which are assigned to the users
@app.route("/FinanceCallTeamMasterData", methods=['GET','POST','OPTIONS'])
@cross_origin(origins="*", headers="*")
def FinanceCallTeamMasterData():
    try:
        res = checkToken(request.headers.get('Authorization'))
        print(res)
        if res:
            cur = conn.cursor()
            cur.execute('''   select policynum as policynumber,policyholder,expirydate,createddate as dateallocated,
  assigneduser as allocatedto from FinanceCallTable where assigneduser is not null and status is null''')
            row_header = [x[0] for x in cur.description]
            records = cur.fetchall()
            data = []
            for record in records:
                data.append(dict(zip(row_header, record)))
            print("Finance Call Team Master data is")
            print(data)

            return jsonify(data), 200
        else:
            return jsonify({"result": "result"}), 401
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in Finance Call Team master data"}), 422


@app.route("/financeteamdailyactivityreport", methods=['GET','POST','OPTIONS'])
@cross_origin(origins="*", headers="*")
def financeteamdailyactivityreport():
    try:
        wb = Workbook()
        ws = wb.active
        ws.freeze_panes = ws['A2']
        ws.cell(1, 1).value = ('Policy Holder')
        ws['A1'].font = Font(bold=True)
        ws.cell(1, 1).border = thin_border
        ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 2).value = ('Policy Number')
        ws['B1'].font = Font(bold=True)
        ws.cell(1, 2).border = thin_border
        ws.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
        ws['B1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 3).value = ('Renewal Date')
        ws['C1'].font = Font(bold=True)
        ws.cell(1, 3).border = thin_border
        ws.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
        ws['C1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 4).value = ('User assigned')
        ws['D1'].font = Font(bold=True)
        ws.cell(1, 4).border = thin_border
        ws.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')
        ws['D1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 5).value = ('Finance review status')
        ws['E1'].font = Font(bold=True)
        ws.cell(1, 5).border = thin_border
        ws.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center')
        ws['E1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 6).value = ('Expected Completion Date')
        ws['F1'].font = Font(bold=True)
        ws.cell(1, 6).border = thin_border
        ws.cell(1, 6).alignment = Alignment(horizontal='center', vertical='center')
        ws['F1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")

        row = 2

        csr = conn.cursor()
        csr.execute("Select policyholder,policynum, case when policynum like 'ICB/PM%' then dateadd(d,1,ExpiryDate) else ExpiryDate end as Renewaldate,assigneduser,status,reviewduedate from finance1  where status = 'completed' and assigneduser is not null and cast(dateofaction as date) = cast(getdate() as date) and policynum in (select distinct policynum from finance1 where ((uwreview = 1) or policynum not in (select c.policynum from finance1 f join uw1 c on c.policynum = f.policynum where c.assigneduser is not null)))")
        rcrds = csr.fetchall()
        print(rcrds)
        col = 1
        AErecords = len(rcrds)
        print(AErecords)
        for r in (rcrds):
            ws.cell(row, col).value = r[0]
            ws.cell(row, col).border = thin_border
            ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 1).value = r[1]
            ws.cell(row, col + 1).border = thin_border
            ws.cell(row, col + 1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 2).value = r[2]
            ws.cell(row, col + 2).border = thin_border
            ws.cell(row, col + 2).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 3).value = r[3]
            ws.cell(row, col + 3).border = thin_border
            ws.cell(row, col + 3).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 4).value = r[4]
            ws.cell(row, col + 4).border = thin_border
            ws.cell(row, col + 4).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 5).value = r[5]
            ws.cell(row, col + 5).border = thin_border
            ws.cell(row, col + 5).alignment = Alignment(horizontal='left', vertical='center')
            row += 1

        for col in ws.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1
            ws.column_dimensions[column].width = adjusted_width

        dt = (datetime.today()).strftime('%d%b%y')
        print(dt)
        file_name = 'c:/easyRenewals/Reports/Temporary/FinanceTeamDailyActivityReport' + '_' + dt + '.xlsx'
        wb.save(file_name)
        print(file_name)
        return send_file(file_name)
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in finance Team Daily Activity Report"}), 422


@app.route("/uwteamdailyactivityreport", methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def uwteamdailyactivityreport():
    try:
        wb = Workbook()
        ws = wb.active
        ws.freeze_panes = ws['A2']
        ws.cell(1, 1).value = ('Policy Holder')
        ws['A1'].font = Font(bold=True)
        ws.cell(1, 1).border = thin_border
        ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 2).value = ('Policy Number')
        ws['B1'].font = Font(bold=True)
        ws.cell(1, 2).border = thin_border
        ws.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
        ws['B1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 3).value = ('Renewal Date')
        ws['C1'].font = Font(bold=True)
        ws.cell(1, 3).border = thin_border
        ws.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
        ws['C1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 4).value = ('User assigned')
        ws['D1'].font = Font(bold=True)
        ws.cell(1, 4).border = thin_border
        ws.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')
        ws['D1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 5).value = ('UW review status')
        ws['E1'].font = Font(bold=True)
        ws.cell(1, 5).border = thin_border
        ws.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center')
        ws['E1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 6).value = ('Expected Completion Date')
        ws['F1'].font = Font(bold=True)
        ws.cell(1, 6).border = thin_border
        ws.cell(1, 6).alignment = Alignment(horizontal='center', vertical='center')
        ws['F1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")

        row = 2

        csr = conn.cursor()
        csr.execute("Select policyholder,policynum, case when policynum like 'ICB/PM%' then dateadd(d,1,ExpiryDate) else ExpiryDate end as Renewaldate,assigneduser,status,reviewduedate from uw1  where status = 'completed'  and assigneduser is not null and cast(dateofaction as date) = cast(getdate() as date)")
        rcrds = csr.fetchall()
        print(rcrds)

        col = 1
        AErecords = len(rcrds)
        print(AErecords)
        for r in (rcrds):
            ws.cell(row, col).value = r[0]
            ws.cell(row, col).border = thin_border
            ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 1).value = r[1]
            ws.cell(row, col + 1).border = thin_border
            ws.cell(row, col + 1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 2).value = r[2]
            ws.cell(row, col + 2).border = thin_border
            ws.cell(row, col + 2).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 3).value = r[3]
            ws.cell(row, col + 3).border = thin_border
            ws.cell(row, col + 3).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 4).value = r[4]
            ws.cell(row, col + 4).border = thin_border
            ws.cell(row, col + 4).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 5).value = r[5]
            ws.cell(row, col + 5).border = thin_border
            ws.cell(row, col + 5).alignment = Alignment(horizontal='left', vertical='center')
            row += 1

        for col in ws.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1
            ws.column_dimensions[column].width = adjusted_width

        dt = (datetime.today()).strftime('%d%b%y')
        print(dt)
        file_name = 'c:/easyRenewals/Reports/Temporary/UWTeamDailyActivityReport' + '_' + dt + '.xlsx'
        wb.save(file_name)
        print(file_name)
        return send_file(file_name)
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in UW Team Daily Activity Report"}), 422

@app.route("/financeteamescalationreport", methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def financeteamescalationreport():
    try:
        wb = Workbook()
        ws = wb.active
        ws.freeze_panes = ws['A2']
        ws.cell(1, 1).value = ('Policy Holder')
        ws['A1'].font = Font(bold=True)
        ws.cell(1, 1).border = thin_border
        ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 2).value = ('Policy Number')
        ws['B1'].font = Font(bold=True)
        ws.cell(1, 2).border = thin_border
        ws.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
        ws['B1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 3).value = ('Renewal Date')
        ws['C1'].font = Font(bold=True)
        ws.cell(1, 3).border = thin_border
        ws.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
        ws['C1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 4).value = ('Review allocation')
        ws['D1'].font = Font(bold=True)
        ws.cell(1, 4).border = thin_border
        ws.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')
        ws['D1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 5).value = ('Days left')
        ws['E1'].font = Font(bold=True)
        ws.cell(1, 5).border = thin_border
        ws.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center')
        ws['E1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 6).value = ('User Assigned')
        ws['F1'].font = Font(bold=True)
        ws.cell(1, 6).border = thin_border
        ws.cell(1, 6).alignment = Alignment(horizontal='center', vertical='center')
        ws['F1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")

        row = 2

        csr = conn.cursor()
        csr.execute("Select policyholder,policynum, case when policynum like 'ICB/PM%' then dateadd(d,1,ExpiryDate) else ExpiryDate end as Renewaldate,Createddate,reviewduein,assigneduser from finance1  where  status is null and reviewduein between 0 and 2  and assigneduser is not null and policynum in (select distinct policynum from finance1 where ((uwreview = 1) or policynum not in (select c.policynum from finance1 f join uw1 c on c.policynum = f.policynum where c.assigneduser is not null)))")
        rcrds = csr.fetchall()
        print(rcrds)

        col = 1

        for r in (rcrds):
            ws.cell(row, col).value = r[0]
            ws.cell(row, col).border = thin_border
            ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 1).value = r[1]
            ws.cell(row, col + 1).border = thin_border
            ws.cell(row, col + 1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 2).value = r[2]
            ws.cell(row, col + 2).border = thin_border
            ws.cell(row, col + 2).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 3).value = r[3]
            ws.cell(row, col + 3).border = thin_border
            ws.cell(row, col + 3).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 4).value = r[4]
            ws.cell(row, col + 4).border = thin_border
            ws.cell(row, col + 4).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 5).value = r[5]
            ws.cell(row, col + 5).border = thin_border
            ws.cell(row, col + 5).alignment = Alignment(horizontal='left', vertical='center')
            row += 1

        for col in ws.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1
            ws.column_dimensions[column].width = adjusted_width

        dt = (datetime.today()).strftime('%d%b%y')
        print(dt)
        file_name = 'c:/easyRenewals/Reports/Temporary/FinanceTeamEscalationReport' + '_' + dt + '.xlsx'
        wb.save(file_name)
        print(file_name)
        return send_file(file_name)
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in Finance Team Escalation Report"}), 422

@app.route("/uwteamescalationreport", methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def uwteamescalationreport():
    try:
        wb = Workbook()
        ws = wb.active
        ws.freeze_panes = ws['A2']
        ws.cell(1, 1).value = ('Policy Holder')
        ws['A1'].font = Font(bold=True)
        ws.cell(1, 1).border = thin_border
        ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 2).value = ('Policy Number')
        ws['B1'].font = Font(bold=True)
        ws.cell(1, 2).border = thin_border
        ws.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
        ws['B1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 3).value = ('Renewal Date')
        ws['C1'].font = Font(bold=True)
        ws.cell(1, 3).border = thin_border
        ws.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
        ws['C1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 4).value = ('Review allocation')
        ws['D1'].font = Font(bold=True)
        ws.cell(1, 4).border = thin_border
        ws.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')
        ws['D1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 5).value = ('Days left')
        ws['E1'].font = Font(bold=True)
        ws.cell(1, 5).border = thin_border
        ws.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center')
        ws['E1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 6).value = ('User Assigned')
        ws['F1'].font = Font(bold=True)
        ws.cell(1, 6).border = thin_border
        ws.cell(1, 6).alignment = Alignment(horizontal='center', vertical='center')
        ws['F1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")

        row = 2

        csr = conn.cursor()
        csr.execute("Select policyholder,policynum, case when policynum like 'ICB/PM%' then dateadd(d,1,ExpiryDate) else ExpiryDate end as Renewaldate,Createddate,reviewduein,assigneduser from uw1  where  status is null and reviewduein between 0 and 2  and assigneduser is not null")
        rcrds = csr.fetchall()
        print(rcrds)

        col = 1

        for r in (rcrds):
            ws.cell(row, col).value = r[0]
            ws.cell(row, col).border = thin_border
            ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 1).value = r[1]
            ws.cell(row, col + 1).border = thin_border
            ws.cell(row, col + 1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 2).value = r[2]
            ws.cell(row, col + 2).border = thin_border
            ws.cell(row, col + 2).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 3).value = r[3]
            ws.cell(row, col + 3).border = thin_border
            ws.cell(row, col + 3).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 4).value = r[4]
            ws.cell(row, col + 4).border = thin_border
            ws.cell(row, col + 4).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 5).value = r[5]
            ws.cell(row, col + 5).border = thin_border
            ws.cell(row, col + 5).alignment = Alignment(horizontal='left', vertical='center')
            row += 1

        for col in ws.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1
            ws.column_dimensions[column].width = adjusted_width

        dt = (datetime.today()).strftime('%d%b%y')
        print(dt)
        file_name = 'c:/easyRenewals/Reports/Temporary/UWTeamEscalationReport' + '_' + dt + '.xlsx'
        wb.save(file_name)
        print(file_name)
        return send_file(file_name)
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in UW Team Escalation Report"}), 422


@app.route("/financeteamoverduereport", methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def financeteamoverduereport():
    try:
        wb = Workbook()
        ws = wb.active
        ws.freeze_panes = ws['A2']
        ws.cell(1, 1).value = ('Policy Holder')
        ws['A1'].font = Font(bold=True)
        ws.cell(1, 1).border = thin_border
        ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 2).value = ('Policy Number')
        ws['B1'].font = Font(bold=True)
        ws.cell(1, 2).border = thin_border
        ws.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
        ws['B1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 3).value = ('Renewal Date')
        ws['C1'].font = Font(bold=True)
        ws.cell(1, 3).border = thin_border
        ws.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
        ws['C1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 4).value = ('Review allocation')
        ws['D1'].font = Font(bold=True)
        ws.cell(1, 4).border = thin_border
        ws.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')
        ws['D1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 5).value = ('Overdue days')
        ws['E1'].font = Font(bold=True)
        ws.cell(1, 5).border = thin_border
        ws.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center')
        ws['E1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 6).value = ('User Assigned')
        ws['F1'].font = Font(bold=True)
        ws.cell(1, 6).border = thin_border
        ws.cell(1, 6).alignment = Alignment(horizontal='center', vertical='center')
        ws['F1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")

        row = 2

        csr = conn.cursor()
        csr.execute("Select policyholder,policynum, case when policynum like 'ICB/PM%' then dateadd(d,1,ExpiryDate) else ExpiryDate end as Renewaldate,Createddate,reviewduein,assigneduser from finance1  where  status is null and reviewduein < 0  and assigneduser is not null and policynum in (select distinct policynum from finance1 where ((uwreview = 1) or policynum not in (select c.policynum from finance1 f join uw1 c on c.policynum = f.policynum where c.assigneduser is not null)))")
        rcrds = csr.fetchall()
        print(rcrds)

        col = 1

        for r in (rcrds):
            ws.cell(row, col).value = r[0]
            ws.cell(row, col).border = thin_border
            ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 1).value = r[1]
            ws.cell(row, col + 1).border = thin_border
            ws.cell(row, col + 1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 2).value = r[2]
            ws.cell(row, col + 2).border = thin_border
            ws.cell(row, col + 2).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 3).value = r[3]
            ws.cell(row, col + 3).border = thin_border
            ws.cell(row, col + 3).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 4).value = r[4]
            ws.cell(row, col + 4).border = thin_border
            ws.cell(row, col + 4).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 5).value = r[5]
            ws.cell(row, col + 5).border = thin_border
            ws.cell(row, col + 5).alignment = Alignment(horizontal='left', vertical='center')
            row += 1

        for col in ws.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1
            ws.column_dimensions[column].width = adjusted_width

        dt = (datetime.today()).strftime('%d%b%y')
        print(dt)
        file_name = 'c:/easyRenewals/Reports/Temporary/FinanceTeamOverdueReport' + '_' + dt + '.xlsx'
        wb.save(file_name)
        print(file_name)
        return send_file(file_name)
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in Finance Team Over due Report"}), 422


@app.route("/uwteamoverduereport", methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def uwteamoverduereport():
    try:
        wb = Workbook()
        ws = wb.active
        ws.freeze_panes = ws['A2']
        ws.cell(1, 1).value = ('Policy Holder')
        ws['A1'].font = Font(bold=True)
        ws.cell(1, 1).border = thin_border
        ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 2).value = ('Policy Number')
        ws['B1'].font = Font(bold=True)
        ws.cell(1, 2).border = thin_border
        ws.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
        ws['B1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 3).value = ('Renewal Date')
        ws['C1'].font = Font(bold=True)
        ws.cell(1, 3).border = thin_border
        ws.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
        ws['C1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 4).value = ('Review allocation')
        ws['D1'].font = Font(bold=True)
        ws.cell(1, 4).border = thin_border
        ws.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')
        ws['D1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 5).value = ('Overdue days')
        ws['E1'].font = Font(bold=True)
        ws.cell(1, 5).border = thin_border
        ws.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center')
        ws['E1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 6).value = ('User Assigned')
        ws['F1'].font = Font(bold=True)
        ws.cell(1, 6).border = thin_border
        ws.cell(1, 6).alignment = Alignment(horizontal='center', vertical='center')
        ws['F1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")

        row = 2

        csr = conn.cursor()
        csr.execute("Select policyholder,policynum, case when policynum like 'ICB/PM%' then dateadd(d,1,ExpiryDate) else ExpiryDate end as Renewaldate,Createddate,reviewduein,assigneduser from uw1  where  status is null and reviewduein < 0  and assigneduser is not null")
        rcrds = csr.fetchall()
        print(rcrds)

        col = 1
        AErecords = len(rcrds)
        print(AErecords)
        for r in (rcrds):
            ws.cell(row, col).value = r[0]
            ws.cell(row, col).border = thin_border
            ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 1).value = r[1]
            ws.cell(row, col + 1).border = thin_border
            ws.cell(row, col + 1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 2).value = r[2]
            ws.cell(row, col + 2).border = thin_border
            ws.cell(row, col + 2).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 3).value = r[3]
            ws.cell(row, col + 3).border = thin_border
            ws.cell(row, col + 3).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 4).value = r[4]
            ws.cell(row, col + 4).border = thin_border
            ws.cell(row, col + 4).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 5).value = r[5]
            ws.cell(row, col + 5).border = thin_border
            ws.cell(row, col + 5).alignment = Alignment(horizontal='left', vertical='center')
            row += 1

        for col in ws.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1
            ws.column_dimensions[column].width = adjusted_width

        dt = (datetime.today()).strftime('%d%b%y')
        print(dt)
        file_name = 'c:/easyRenewals/Reports/Temporary/UWTeamOverdueReport' + '_' + dt + '.xlsx'
        wb.save(file_name)
        print(file_name)
        return send_file(file_name)
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in UW Team Over due Report"}), 422

@app.route("/UWTeamRejectionReport", methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def UWTeamRejectionReport():
    try:
        wb = Workbook()
        ws = wb.active
        ws.freeze_panes = ws['A2']
        ws.cell(1, 1).value = ('Policy Number')
        ws['A1'].font = Font(bold=True)
        ws.cell(1, 1).border = thin_border
        ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 2).value = ('Policy Expiry Date')
        ws['B1'].font = Font(bold=True)
        ws.cell(1, 2).border = thin_border
        ws.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
        ws['B1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 3).value = ('Rejected Date')
        ws['C1'].font = Font(bold=True)
        ws.cell(1, 3).border = thin_border
        ws.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
        ws['C1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 4).value = ('Rejected By')
        ws['D1'].font = Font(bold=True)
        ws.cell(1, 4).border = thin_border
        ws.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')
        ws['D1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 5).value = ('Rejection Reason')
        ws['E1'].font = Font(bold=True)
        ws.cell(1, 5).border = thin_border
        ws.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center')
        ws['E1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")

        today = datetime.today()
        last_monday = today - timedelta(days=today.weekday())
        sunday = today + timedelta(days=-today.weekday()-1, weeks=1)

        row = 2

        csr = conn.cursor()
        csr.execute("Select ltrim(rtrim(policynum)), expirydate,cast(Dateofaction as date) as rejecteddate,username,note from uw1  where  status like 'Completed' and TypeofAction like 'Reject' and cast(dateofaction as date) between ? and ?",(last_monday,sunday))
        rcrds = csr.fetchall()
        print(rcrds)

        col = 1
        AErecords = len(rcrds)
        print(AErecords)
        for r in (rcrds):
            ws.cell(row, col).value = r[0]
            ws.cell(row, col).border = thin_border
            ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 1).value = r[1]
            ws.cell(row, col + 1).border = thin_border
            ws.cell(row, col + 1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 2).value = r[2]
            ws.cell(row, col + 2).border = thin_border
            ws.cell(row, col + 2).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 3).value = r[3]
            ws.cell(row, col + 3).border = thin_border
            ws.cell(row, col + 3).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 4).value = r[4]
            ws.cell(row, col + 4).border = thin_border
            ws.cell(row, col + 4).alignment = Alignment(horizontal='left', vertical='center')

            row += 1

        for col in ws.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1
            ws.column_dimensions[column].width = adjusted_width

        dt = (datetime.today()).strftime('%d%b%y')
        print(dt)
        file_name = 'c:/easyRenewals/Reports/Temporary/UWTeamRejectionReport' + '_' + dt + '.xlsx'
        wb.save(file_name)
        print(file_name)
        return send_file(file_name)
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in UW Team Rejection Report"}), 422


@app.route("/financeteamweeklyreport", methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def financeteamweeklyreport():
    try:
        list_of_files = glob.glob('C:/easyRenewals/Reports/financeTeamWeeklyReports/*.xlsx')
        file_name = max(list_of_files, key=os.path.getctime)
        print(file_name)
        return send_file(file_name)
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in Finance Team Weekly Report"}), 422

@app.route("/uwteamweeklyreport", methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def uwteamweeklyreport():
    try:
        list_of_files = glob.glob('C:/easyRenewals/Reports/UWTeamWeeklyReports/*.xlsx')
        print(list_of_files)
        file_name = max(list_of_files, key=os.path.getctime)
        print(file_name)
        return send_file(file_name)
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in UW Team weekly Report"}), 422

@app.route("/financeuserweeklyperformancereport", methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def financeuserweeklyperformancereport():
    try:
        list_of_files = glob.glob('C:/easyRenewals/Reports/financeTeamWeeklyperformanceReports/*.xlsx')
        file_name = max(list_of_files, key=os.path.getctime)
        print(file_name)
        return send_file(file_name)
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in Finance User Weekly Performance Report"}), 422

@app.route("/uwuserweeklyperformancereport", methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def uwuserweeklyperformancereport():
    try:
        list_of_files = glob.glob('C:/easyRenewals/Reports/uwTeamWeeklyperformanceReports/*.xlsx')
        file_name = max(list_of_files, key=os.path.getctime)
        print(file_name)
        return send_file(file_name)
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in uw user weekly performance report"}), 422


#Finance Team Call Data
@app.route('/tempfinancecalldata', methods=['GET','POST','OPTIONS'])
@cross_origin(origins="*", headers="*")
def tempfinancecalldata():
    try:
        cur = conn.cursor()
        res = checkToken(request.headers.get('Authorization'))
        print(res)
        result = False
        if res:
            print("in finiance call data")
            data = request.json
            print(data)
            datas = request.headers.get('Authorization')
            print(datas)
            datas = datas.replace('"', '')
            print(datas)
            decoded = jwt.decode(datas, str('secret'), 'utf-8')
            user = decoded.get("Name")
            print(user)
            permissions = decoded.get("permissions")
            print(permissions)
            try:
                cur.execute("select * from financecallTable where assigneduser = ?",(user))
                row_headers = [x[0] for x in cur.description]
                records = cur.fetchall()
                json_data = []
                for record in records:
                    json_data.append(dict(zip(row_headers, record)))
                print("Data is")
                print(json_data)
                result = True
                return jsonify(json_data)
            except jwt.ExpiredSignatureError:
                print("sign")
                result = False
            if result:
                return jsonify({"result": result}), 200
            else:
                return jsonify({"result": result}), 401
        else:
            result = False
            return jsonify({"result": result}), 401
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in Finance Team Call data"}), 422

#Finance Call Team Action

@app.route("/fincallteamaction", methods=['GET','POST','OPTIONS'])
@cross_origin(origins="*", headers="*")
def fincallteamaction():
    try:
        res = checkToken(request.headers.get('Authorization'))
        print(res)
        result = False
        if res:
            Data = request.get_json()
            print(Data)
            polnum = Data.get('PolicyNum')
            polnum = polnum.strip()
            print(polnum)
            status = Data.get('status')
            username = Data.get('name')
            dateofaction = datetime.today()

            cur = conn.cursor()
            cur.execute(
                "update financecalltable set dateofaction = ?,username= ?,status = ?,notes = ? where policynum = ?",
                (dateofaction, username, status, Data.get('note'), polnum))
            cur.commit()

            return jsonify({"result":"success"}), 200

        else:
            return jsonify({"result": result}), 401
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in Finance Team Action"}), 422


@app.route("/FinanceCallReport", methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def FinanceCallReport():
    try:
        wb = Workbook()
        ws = wb.active
        ws.freeze_panes = ws['A2']
        ws.cell(1, 1).value = ('Policy Number')
        ws['A1'].font = Font(bold=True)
        ws.cell(1, 1).border = thin_border
        ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 2).value = ('Premium Due')
        ws['B1'].font = Font(bold=True)
        ws.cell(1, 2).border = thin_border
        ws.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
        ws['B1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 3).value = ('Policy Expiry Date')
        ws['C1'].font = Font(bold=True)
        ws.cell(1, 3).border = thin_border
        ws.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
        ws['C1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 4).value = ('Date Allocated')
        ws['D1'].font = Font(bold=True)
        ws.cell(1, 4).border = thin_border
        ws.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')
        ws['D1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 5).value = ('Date Followed up')
        ws['E1'].font = Font(bold=True)
        ws.cell(1, 5).border = thin_border
        ws.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center')
        ws['E1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 6).value = ('Followed up by')
        ws['F1'].font = Font(bold=True)
        ws.cell(1, 6).border = thin_border
        ws.cell(1, 6).alignment = Alignment(horizontal='center', vertical='center')
        ws['F1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 7).value = ('No of Days for Completion')
        ws['G1'].font = Font(bold=True)
        ws.cell(1, 7).border = thin_border
        ws.cell(1, 7).alignment = Alignment(horizontal='center', vertical='center')
        ws['G1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 8).value = ('Notes')
        ws['H1'].font = Font(bold=True)
        ws.cell(1, 8).border = thin_border
        ws.cell(1, 8).alignment = Alignment(horizontal='center', vertical='center')
        ws['H1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")

        row = 2

        csr = conn.cursor()
        csr.execute("select policynum,premiumdue,expirydate,createddate,cast(dateofaction as date) dateofaction,username,datediff(d,createddate,cast(dateofaction as date)) as dayscompletion,notes from financecalltable where status like 'Completed'")
        rcrds = csr.fetchall()
        print(rcrds)

        col = 1

        for r in (rcrds):
            ws.cell(row, col).value = r[0]
            ws.cell(row, col).border = thin_border
            ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 1).number_format = '0.00'
            ws.cell(row, col + 1).value = float(r[1])
            ws.cell(row, col + 1).border = thin_border
            ws.cell(row, col + 1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 2).value = r[2]
            ws.cell(row, col + 2).border = thin_border
            ws.cell(row, col + 2).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 3).value = r[3]
            ws.cell(row, col + 3).border = thin_border
            ws.cell(row, col + 3).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 4).value = r[4]
            ws.cell(row, col + 4).border = thin_border
            ws.cell(row, col + 4).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 5).value = r[5]
            ws.cell(row, col + 5).border = thin_border
            ws.cell(row, col + 5).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 6).value = r[6]
            ws.cell(row, col + 6).border = thin_border
            ws.cell(row, col + 6).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row, col + 7).value = r[7]
            ws.cell(row, col + 7).border = thin_border
            ws.cell(row, col + 7).alignment = Alignment(horizontal='left', vertical='center')

            row += 1

        for col in ws.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1
            ws.column_dimensions[column].width = adjusted_width

        dt = (datetime.today()).strftime('%d%b%y')
        print(dt)
        file_name = 'c:/easyRenewals/Reports/Temporary/FinanceCallReport' + '_' + dt + '.xlsx'
        wb.save(file_name)
        print(file_name)
        return send_file(file_name)
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in Finance Team Call report"}), 422
    

#needs to be updated
@app.route("/EmailNotificationsReport", methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def EmailNotificationsReport():
    try:
        dt = (datetime.today() - timedelta(days=1)).strftime('%Y-%m-%d')
        print(dt)

        file_name = 'c:/Mantra/EasyRenewals/Reports/'+str((datetime.today() - timedelta(days=1)).strftime('%B'))+'/EmailFollowUps_'+dt+'.xls'
        print(file_name)
        return send_file(file_name)
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in uw user weekly performance report"}), 422

#Needs to be updated
@app.route("/SMSNotificationsReport", methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def SMSNotificationsReport():
    try:
        dt = (datetime.today() - timedelta(days=1)).strftime('%Y-%m-%d')
        print(dt)
        file_name = 'c:/Mantra/EasyRenewals/Reports/sms/'+str((datetime.today() - timedelta(days=1)).strftime('%B'))+'/smsfollowups_'+dt+'.xls'
        #file_name = max(list_of_files, key=os.path.getctime)
        print(file_name)
        return send_file(file_name)
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in uw user weekly performance report"}), 422


#Needs to be updated on 24Nov2021
@app.route("/MonthlyStatsReport", methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def MonthlyStatsReport():
    try:
        dt = (datetime.today() - timedelta(days=1)).strftime('%Y-%m-%d')
        print(dt)
        file_name = 'c:/Mantra/PythonReportsRenewals/Monthly_Overdue/MonthlyStats_'+dt+'.xlsx'
        #file_name = max(list_of_files, key=os.path.getctime)
        print(file_name)
        return send_file(file_name)
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in MonthlyStats Report"}), 422
 
#Needs to be updated on 24Nov2021
@app.route("/MonthlyEmailSMSNotificationsReport", methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="*", headers="*")
def MonthlyEmailSMSNotificationsReport():
    try:
        dt = (datetime.today() - timedelta(days=1)).strftime('%Y-%m-%d')
        print(dt)
        file_name = 'c:/Mantra/PythonReportsRenewals/Monthly_EmailSms/MonthlyEmailSMSReport_'+dt+'.xlsx'
        #file_name = max(list_of_files, key=os.path.getctime)
        print(file_name)
        return send_file(file_name)
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in Monthly Email SMS Notifications Report"}), 422




@app.route("/FinanceCallTeamPolicyaAssignment", methods=['GET','POST','OPTIONS'])
@cross_origin(origins="*", headers="*")
def FinanceCallTeamPolicyaAssignment():
    try:
        res = checkToken(request.headers.get('Authorization'))
        print(res)
        if res:
            Data = request.get_json()
            print(Data)
            polnum = Data.get('Policynumbers')
            print(polnum)
            allocatedto = Data['values']['username']
            print(allocatedto)
            allocateddate = datetime.today()
            allocatedby = Data.get('username')
            print(allocatedby)
            for policynum in polnum:
                cur = conn.cursor()
                cur.execute("select assigneduser from FinanceCallTable where policynum like ?",policynum)
                r = cur.fetchone()
                user =None
                for i in r:
                    user = r[0]
                print(user)
                cur.execute("update uw1 set previouslyallocatedto = ? where policynum like ?",(user,policynum))
                cur.commit()
                cur.execute(
                    "update FinanceCallTable set allocateddate = ?,assigneduser= ?, allocatedby = ?  where policynum like ?",
                    (allocateddate, allocatedto, allocatedby, policynum))
                cur.commit()

            return jsonify({"result":"success"}), 200

        else:
            return jsonify({"result": "result"}), 401
    except Exception as e:
        logging.warning(datetime.now())
        logging.exception("error")
        return jsonify({"message": "Error occured in Finance Call Team policy Assignment function"}), 422

if __name__ == '__main__':
    app.secret_key = 'random string'
    app.debug = True
    app.run(host='172.21.32.34', port=5001)





